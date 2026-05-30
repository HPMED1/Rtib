"""Exporter round-trips. All three formats keep nulls and parse_error rows."""

from __future__ import annotations

import csv
import json

import pytest
from openpyxl import load_workbook

from rtib.core.exporters import _coerce_xlsx_value, export_csv, export_json, export_xlsx
from rtib.core.pipeline import RowResult
from rtib.core.schema import Header


@pytest.fixture
def headers():
    return [
        Header(name="Title", description="movie title"),
        Header(name="Year", description="release year"),
    ]


@pytest.fixture
def results():
    return [
        RowResult(raw="The Matrix 1999.mkv", values={"title": "The Matrix", "year": "1999"}),
        RowResult(raw="Inception 2010.mp4", values={"title": "Inception", "year": "2010"}),
        # Null field — model couldn't determine the year
        RowResult(raw="Mystery.unknown", values={"title": "Mystery", "year": None}),
        # parse_error row — model failed entirely
        RowResult(raw="garbled garbage", values=None, error="invalid json"),
    ]


class TestJsonExport:
    def test_basic(self, tmp_path, headers, results):
        path = tmp_path / "out.json"
        export_json(path, headers, results)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert len(data) == 4
        assert data[0] == {"title": "The Matrix", "year": "1999"}
        assert data[2] == {"title": "Mystery", "year": None}
        # parse_error row has nulls for every field by default
        assert data[3] == {"title": None, "year": None}

    def test_with_status(self, tmp_path, headers, results):
        path = tmp_path / "out.json"
        export_json(path, headers, results, include_status=True)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data[0]["_status"] == "ok"
        assert data[2]["_status"] == "ok"  # null field is still parsed ok
        assert data[3]["_status"] == "parse_error"

    def test_with_raw(self, tmp_path, headers, results):
        path = tmp_path / "out.json"
        export_json(path, headers, results, include_raw=True)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data[0]["_raw"] == "The Matrix 1999.mkv"
        assert data[3]["_raw"] == "garbled garbage"


class TestCsvExport:
    def test_basic(self, tmp_path, headers, results):
        path = tmp_path / "out.csv"
        export_csv(path, headers, results)
        with path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 4
        assert rows[0]["title"] == "The Matrix"
        assert rows[0]["year"] == "1999"
        # Null becomes empty string in CSV
        assert rows[2]["year"] == ""
        # parse_error row has empty cells
        assert rows[3]["title"] == ""

    def test_with_status(self, tmp_path, headers, results):
        path = tmp_path / "out.csv"
        export_csv(path, headers, results, include_status=True)
        with path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["_status"] == "ok"
        assert rows[3]["_status"] == "parse_error"

    def test_handles_commas_in_values(self, tmp_path, headers):
        path = tmp_path / "out.csv"
        results = [
            RowResult(raw="x", values={"title": "Romeo, Juliet", "year": "1996"}),
        ]
        export_csv(path, headers, results)
        with path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["title"] == "Romeo, Juliet"

    def test_handles_newlines_in_values(self, tmp_path, headers):
        path = tmp_path / "out.csv"
        results = [
            RowResult(raw="x", values={"title": "Two\nLines", "year": "2000"}),
        ]
        export_csv(path, headers, results)
        with path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["title"] == "Two\nLines"


class TestXlsxExport:
    def test_basic(self, tmp_path, headers, results):
        path = tmp_path / "out.xlsx"
        export_xlsx(path, headers, results)
        wb = load_workbook(path)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        assert rows_data[0] == ("Title", "Year")  # uses display name, not slug
        # Year coerces from "1999" -> int 1999 (type-aware XLSX).
        assert rows_data[1] == ("The Matrix", 1999)
        assert rows_data[3] == ("Mystery", None)
        # parse_error rows are written as empty strings; openpyxl normalizes
        # them back to None on read but the row still exists in the sheet,
        # which is the behaviour Excel users actually see.
        assert len(rows_data) == 5  # header + 3 data + parse_error row

    def test_with_status_column(self, tmp_path, headers, results):
        path = tmp_path / "out.xlsx"
        export_xlsx(path, headers, results, include_status=True)
        wb = load_workbook(path)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        assert rows_data[0] == ("Title", "Year", "_status")
        assert rows_data[1][2] == "ok"
        assert rows_data[4][2] == "parse_error"

    def test_sheet_named_rtib(self, tmp_path, headers, results):
        path = tmp_path / "out.xlsx"
        export_xlsx(path, headers, results)
        wb = load_workbook(path)
        assert wb.active.title == "Rtib"


class TestXlsxTypeCoercion:
    """Per-cell coercion: years as ints, dates as dates, identifiers as text."""

    def test_int_year(self):
        assert _coerce_xlsx_value("1999") == 1999
        assert isinstance(_coerce_xlsx_value("1999"), int)

    def test_zero(self):
        assert _coerce_xlsx_value("0") == 0

    def test_negative_int(self):
        assert _coerce_xlsx_value("-42") == -42

    def test_leading_zero_stays_string(self):
        # Looks like a zero-padded ID — keep it intact.
        assert _coerce_xlsx_value("01234") == "01234"

    def test_long_number_stays_string(self):
        # 11 digits — likely a phone number or some external ID.
        assert _coerce_xlsx_value("12345678901") == "12345678901"

    def test_float(self):
        assert _coerce_xlsx_value("12.5") == 12.5
        assert isinstance(_coerce_xlsx_value("12.5"), float)

    def test_negative_float(self):
        assert _coerce_xlsx_value("-3.14") == -3.14

    def test_iso_date(self):
        from datetime import date as _date

        assert _coerce_xlsx_value("2024-01-15") == _date(2024, 1, 15)

    def test_iso_datetime(self):
        from datetime import datetime as _dt

        assert _coerce_xlsx_value("2024-01-15T10:30:00") == _dt(2024, 1, 15, 10, 30)

    def test_resolution_stays_string(self):
        # The classic gotcha — "1080p" starts with digits but isn't a number.
        assert _coerce_xlsx_value("1080p") == "1080p"

    def test_phone_stays_string(self):
        assert _coerce_xlsx_value("555-1234") == "555-1234"

    def test_empty_returns_none(self):
        assert _coerce_xlsx_value("") is None
        assert _coerce_xlsx_value(None) is None

    def test_string_with_internal_separator(self):
        assert _coerce_xlsx_value("1,000") == "1,000"  # comma — not a float

    def test_exporter_writes_int_year(self, tmp_path):
        from openpyxl import load_workbook

        headers = [Header(name="Title"), Header(name="Year")]
        results = [
            RowResult(raw="x", values={"title": "The Matrix", "year": "1999"}),
        ]
        path = tmp_path / "out.xlsx"
        export_xlsx(path, headers, results)
        wb = load_workbook(path)
        ws = wb.active
        # Header row + 1 data row
        cell = ws.cell(row=2, column=2)
        assert cell.value == 1999
        assert isinstance(cell.value, int)


class TestEmptyResults:
    def test_json_empty(self, tmp_path, headers):
        path = tmp_path / "out.json"
        export_json(path, headers, [])
        assert json.loads(path.read_text(encoding="utf-8")) == []

    def test_csv_empty_writes_header_only(self, tmp_path, headers):
        path = tmp_path / "out.csv"
        export_csv(path, headers, [])
        with path.open(encoding="utf-8") as f:
            lines = f.read().strip().splitlines()
        assert len(lines) == 1  # header row only
        assert lines[0] == "title,year"
